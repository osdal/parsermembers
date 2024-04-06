import pandas as pd
from openpyxl import load_workbook

# Чтение данных из файла final.xlsx с использованием openpyxl
wb = load_workbook('final.xlsx')
ws = wb.active

# Создание DataFrame из содержимого файла Excel
data = []
for row in ws.iter_rows(values_only=True):
    data.append(row)
df = pd.DataFrame(data, columns=[col[0] for col in ws.iter_cols(min_row=1, max_row=1, values_only=True)])

# Объединение ячеек в столбце ID, содержащих одинаковые значения
prev_value = None
start_index = None
for idx, value in enumerate(df['ID']):
    if value != prev_value:
        if start_index is not None:
            end_index = idx - 1
            if start_index != end_index:
                ws.merge_cells(start_row=start_index + 1, start_column=1, end_row=end_index + 1, end_column=1)
        start_index = idx
    prev_value = value

# Объединение ячеек в столбцах name и surname, содержащих одинаковые значения
for column in ['name', 'surname']:
    prev_value = None
    start_index = None
    for idx, value in enumerate(df[column]):
        if value != prev_value:
            if start_index is not None:
                end_index = idx - 1
                if start_index != end_index:
                    ws.merge_cells(start_row=start_index + 1, start_column=df.columns.get_loc(column) + 1, end_row=end_index + 1, end_column=df.columns.get_loc(column) + 1)
            start_index = idx
        prev_value = value

# Сохранение изменений в файле final.xlsx
wb.save('final.xlsx')
